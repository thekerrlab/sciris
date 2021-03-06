"""
fileio.py -- code for file management in Sciris 
    
Last update: 5/31/18 (gchadder3)
"""

##############################################################################
### Imports
##############################################################################

# Basic imports
import io
import os
import re
import pickle
import dill
import types
import numpy as np
from glob import glob
from gzip import GzipFile
from contextlib import closing
from . import sc_utils as ut
from .sc_odict import odict
from .sc_dataframe import dataframe

# Handle types and Python 2/3 compatibility
import six
_stringtype = six.string_types[0]
if six.PY3: # Python 3
    from io import BytesIO as IO
    import pickle as pkl
    import copyreg as cpreg
else: # Python 2
    from cStringIO import StringIO as IO
    import cPickle as pkl
    import copy_reg as cpreg




##############################################################################
### Pickling functions
##############################################################################

__all__ = ['loadobj', 'loadstr', 'saveobj', 'dumpstr']


def loadobj(filename=None, folder=None, verbose=True, die=None):
    '''
    Load a saved file.

    Usage:
        obj = loadobj('myfile.obj')
    '''
    
    # Handle loading of either filename or file object
    if isinstance(filename, _stringtype): 
        argtype = 'filename'
        filename = makefilepath(filename=filename, folder=folder) # If it is a file, validate the folder
    else: 
        argtype = 'fileobj'
    kwargs = {'mode': 'rb', argtype: filename}
    with GzipFile(**kwargs) as fileobj:
        filestr = fileobj.read() # Convert it to a string
        obj = unpickler(filestr, die=die) # Actually load it
    if verbose: print('Object loaded from "%s"' % filename)
    return obj


def loadstr(string=None, die=None):
    with closing(IO(string)) as output: # Open a "fake file" with the Gzip string pickle in it.
        with GzipFile(fileobj=output, mode='rb') as fileobj: # Set a Gzip reader to pull from the "file."
            picklestring = fileobj.read() # Read the string pickle from the "file" (applying Gzip decompression).
    obj = unpickler(picklestring, die=die) # Return the object gotten from the string pickle.   
    return obj


def saveobj(filename=None, obj=None, compresslevel=5, verbose=True, folder=None, method='pickle'):
    '''
    Save an object to file -- use compression 5 by default, since more is much slower but not much smaller.
    Once saved, can be loaded with loadobj() (q.v.).

    Usage:
        myobj = ['this', 'is', 'a', 'weird', {'object':44}]
        saveobj('myfile.obj', myobj)
    '''
    
    fullpath = makefilepath(filename=filename, folder=folder, sanitize=True)
    with GzipFile(fullpath, 'wb', compresslevel=compresslevel) as fileobj:
        if method == 'dill': # If dill is requested, use that
            savedill(fileobj, obj)
        else: # Otherwise, try pickle
            try:    savepickle(fileobj, obj) # Use pickle
            except: savedill(fileobj, obj) # ...but use Dill if that fails
        
    if verbose: print('Object saved to "%s"' % fullpath)
    return fullpath


def dumpstr(obj=None):
    with closing(IO()) as output: # Open a "fake file."
        with GzipFile(fileobj=output, mode='wb') as fileobj:  # Open a Gzip-compressing way to write to this "file."
            try:    savepickle(fileobj, obj) # Use pickle
            except: savedill(fileobj, obj) # ...but use Dill if that fails
        output.seek(0) # Move the mark to the beginning of the "file."
        result = output.read() # Read all of the content into result.
    return result




##############################################################################
### Other file functions
##############################################################################

__all__ += ['loadtext', 'savetext', 'getfilelist', 'sanitizefilename', 'makefilepath']


    
def loadtext(filename=None, splitlines=False):
    ''' Convenience function for reading a text file '''
    with open(filename) as f: output = f.read()
    if splitlines: output = output.splitlines()
    return output



def savetext(filename=None, string=None):
    ''' Convenience function for saving a text file -- accepts a string or list of strings '''
    if isinstance(string, list): string = '\n'.join(string) # Convert from list to string)
    if not ut.isstring(string):
        string = str(string)
    with open(filename, 'w') as f: f.write(string)
    return None



def getfilelist(folder=None, ext=None, pattern=None):
    ''' A short-hand since glob is annoying '''
    if folder is None: folder = os.getcwd()
    if pattern is None:
        if ext is None: ext = '*'
        pattern = '*.'+ext
    filelist = sorted(glob(os.path.join(folder, pattern)))
    return filelist



def sanitizefilename(rawfilename):
    '''
    Takes a potentially Linux- and Windows-unfriendly candidate file name, and 
    returns a "sanitized" version that is more usable.
    '''
    filtername = re.sub('[\!\?\"\'<>]', '', rawfilename) # Erase certain characters we don't want at all: !, ?, ", ', <, >
    filtername = re.sub('[:/\\\*\|,]', '_', filtername) # Change certain characters that might be being used as separators from what they were to underscores: space, :, /, \, *, |, comma
    return filtername # Return the sanitized file name.



def makefilepath(filename=None, folder=None, ext=None, default=None, split=False, abspath=True, makedirs=True, verbose=False, sanitize=False):
    '''
    Utility for taking a filename and folder -- or not -- and generating a valid path from them.
    
    Inputs:
        filename = the filename, or full file path, to save to -- in which case this utility does nothing
        folder = the name of the folder to be prepended to the filename
        ext = the extension to ensure the file has
        default = a name or list of names to use if filename is None
        split = whether to return the path and filename separately
        makedirs = whether or not to make the folders to save into if they don't exist
        verbose = how much detail to print
    
    Example:
        makefilepath(filename=None, folder='./congee', ext='prj', default=[project.filename, project.name], split=True, abspath=True, makedirs=True)
    
    Assuming project.filename is None and project.name is "soggyrice" and ./congee doesn't exist:
        * Makes folder ./congee
        * Returns e.g. ('/home/optima/congee', 'soggyrice.prj')
    
    Actual code example from project.py:
        fullpath = makefilepath(filename=filename, folder=folder, default=[self.filename, self.name], ext='prj')
    
    Version: 2017apr04    
    '''
    
    # Initialize
    filefolder = '' # The folder the file will be located in
    filebasename = '' # The filename
    
    # Process filename
    if filename is None:
        defaultnames = ut.promotetolist(default) # Loop over list of default names
        for defaultname in defaultnames:
            if not filename and defaultname: filename = defaultname # Replace empty name with default name
    if filename is not None: # If filename exists by now, use it
        filebasename = os.path.basename(filename)
        filefolder = os.path.dirname(filename)
    if not filebasename: filebasename = 'default' # If all else fails
    
    # Add extension if it's defined but missing from the filebasename
    if ext and not filebasename.endswith(ext): 
        filebasename += '.'+ext
    if verbose:
        print('From filename="%s", default="%s", extension="%s", made basename "%s"' % (filename, default, ext, filebasename))
    
    # Sanitize base filename
    if sanitize: filebasename = sanitizefilename(filebasename)
    
    # Process folder
    if folder is not None: # Replace with specified folder, if defined
        filefolder = folder 
    if abspath: # Convert to absolute path
        filefolder = os.path.abspath(filefolder) 
    if makedirs: # Make sure folder exists
        try: os.makedirs(filefolder)
        except: pass
    if verbose:
        print('From filename="%s", folder="%s", abspath="%s", makedirs="%s", made folder name "%s"' % (filename, folder, abspath, makedirs, filefolder))
    
    fullfile = os.path.join(filefolder, filebasename) # And the full thing
    
    if split: return filefolder, filebasename
    else:     return fullfile # Or can do os.path.split() on output




##############################################################################
### Spreadsheet functions
##############################################################################

__all__ += ['Blobject', 'Spreadsheet', 'loadspreadsheet', 'savespreadsheet']


class Blobject(object):
    ''' A wrapper for a binary file '''

    def __init__(self, source=None, name=None, filename=None):
        # "source" is a specification of where to get the data from
        # It can be anything supported by Blobject.load() which are
        # - A filename, which will get loaded
        # - A io.BytesIO which will get dumped into this instance
    
        # Handle inputs
        if source   is None and filename is not None: source   = filename # Reset the source to be the filename, e.g. Spreadsheet(filename='foo.xlsx')
        if filename is None and ut.isstring(source):  filename = source   # Reset the filename to be the source, e.g. Spreadsheet('foo.xlsx')
        if name     is None and filename is not None: name     = os.path.basename(filename) # If not supplied, use the filename
        
        # Define quantities
        self.name      = name # Name of the object
        self.filename  = filename # Filename (used as default for load/save)
        self.created   = ut.now() # When the object was created
        self.modified  = ut.now() # When it was last modified
        self.blob  = None # The binary data
        self.bytes = None # The filestream representation of the binary data
        if source is not None: self.load(source)
        return None

    def __repr__(self):
        return ut.prepr(self, skip=['blob','bytes'])

    def load(self, source=None):
        '''
        This function loads the spreadsheet from a file or object. If no input argument is supplied,
        then it will read self.bytes, assuming it exists.
        '''
        def read_bin(source):
            ''' Helper to read a binary stream '''
            source.flush()
            source.seek(0)
            output = source.read()
            return output
        
        def read_file(filename):
            ''' Helper to read an actual file '''
            filepath = makefilepath(filename=filename)
            self.filename = filename
            with open(filepath, mode='rb') as f:
                output = f.read()
            return output
            
        if source is None:
            if self.bytes is not None:
                self.blob = read_bin(self.bytes)
                self.bytes = None # Once read in, delete
            else:
                if self.filename is not None:
                    self.blob = read_file(self.filename)
                else:
                    print('Nothing to load: no source or filename supplied and self.bytes is empty.')
        else:
            if isinstance(source,io.BytesIO):
                self.blob = read_bin(source)
            elif ut.isstring(source):
                self.blob = read_file(source)
            else:
                errormsg = 'Input source must be type string (for a filename) or BytesIO, not %s' % type(source)
                raise Exception(errormsg)
        
        self.modified = ut.now()
        return None

    def save(self, filename=None):
        ''' This function writes the spreadsheet to a file on disk. '''
        if filename is None:
            if self.filename is not None:
                filename = self.filename
            elif self.name is not None:
                if self.name.endswith('.xlsx'):
                    filename = self.name
                else:
                    filename = self.name + '.xlsx'
            else:
                filename = 'spreadsheet.xlsx' # Come up with a terrible default name
        filepath = makefilepath(filename=filename)
        with open(filepath, mode='wb') as f:
            f.write(self.blob)
        self.filename = filename
        print('Spreadsheet saved to %s.' % filepath)
        return filepath

    def tofile(self, output=True):
        '''
        Return a file-like object with the contents of the file.
        This can then be used to open the workbook from memory without writing anything to disk e.g.
        - book = openpyxl.load_workbook(self.tofile())
        - book = xlrd.open_workbook(file_contents=self.tofile().read())
        '''
        bytesblob = io.BytesIO(self.blob)
        if output:
            return bytesblob
        else:
            self.bytes = bytesblob
            return None
    
    
    
class Spreadsheet(Blobject):
    '''
    A class for reading and writing Excel files in binary format.No disk IO needs 
    to happen to manipulate the spreadsheets with openpyxl (or xlrd or pandas).

    Version: 2018sep03
    '''
    
    def xlrd(self):
        ''' Return a book as opened by xlrd '''
        import xlrd # Optional import
        book = xlrd.open_workbook(file_contents=self.tofile().read())
        return book
    
    def openpyxl(self):
        ''' Return a book as opened by openpyxl '''
        import openpyxl # Optional iport
        self.tofile(output=False)
        book = openpyxl.load_workbook(self.bytes) # This stream can be passed straight to openpyxl
        return book
        
    def pandas(self):
        ''' Return a book as opened by pandas '''
        import pandas # Optional import
        self.tofile(output=False)
        book = pandas.ExcelFile(self.bytes)
        return book
    
    def update(self, book):
        self.tofile(output=False)
        book.save(self.bytes)
        self.load()
        return None
    
    @staticmethod
    def _getsheet(book, sheetname=None, sheetnum=None):
        if   sheetname is not None: sheet = book[sheetname]
        elif sheetnum  is not None: sheet = book[book.sheetnames[sheetnum]]
        else:                       sheet = book.active
        return sheet
    
    def readcells(self, *args, **kwargs):
        ''' Alias to loadspreadsheet() '''
        if 'method' in kwargs:
            method = kwargs['method']
            kwargs.pop('method')
        else:
            method = None
        if method is None: method = 'xlrd'
        f = self.tofile()
        kwargs['fileobj'] = f
        if method == 'xlrd':
            output = loadspreadsheet(*args, **kwargs)
        elif method == 'openpyxl':
            book = self.openpyxl()
            ws = self._getsheet(book=book, sheetname=kwargs.get('sheetname'), sheetnum=kwargs.get('sheetname'))
            rawdata = tuple(ws.rows)
            output = np.empty(np.shape(rawdata), dtype=object)
            for r,rowdata in enumerate(rawdata):
                for c,val in enumerate(rowdata):
                    output[r][c] = rawdata[r][c].value
        else:
            errormsg = 'Reading method not found; must be one of xlrd, openpyxl, or pandas, not %s' % method
            raise Exception(errormsg)
        return output
    
    def writecells(self, cells=None, row=None, col=None, vals=None, sheetname=None, sheetnum=None, verbose=False):
        '''
        Specify cells to write. Can supply either a list of cells of the same length
        as the values, or else specify a starting row and column and write the values
        from there.
        '''
        import openpyxl # Optional import
        
        # Load workbook
        self.tofile(output=False) # Convert to bytes
        wb = openpyxl.load_workbook(self.bytes)
        if verbose: print('Workbook loaded: %s' % wb)
        
        # Get right worksheet
        ws = self._getsheet(book=wb, sheetname=sheetname, sheetnum=sheetnum)
        if verbose: print('Worksheet loaded: %s' % ws)
        
        # Determine the cells
        if cells is not None: # A list of cells is supplied
            cells = ut.promotetolist(cells)
            vals  = ut.promotetolist(vals)
            if len(cells) != len(vals):
                errormsg = 'If using cells, cells and vals must have the same length (%s vs. %s)' % (len(cells), len(vals))
                raise Exception(errormsg)
            for cell,val in zip(cells,vals):
                try:
                    ws[cell] = val
                    if verbose: print('  Cell %s = %s' % (cell,val))
                except Exception as E:
                    errormsg = 'Could not write "%s" to cell "%s": %s' % (val, cell, repr(E))
                    raise Exception(errormsg)
        else:# Cells aren't supplied, assume a matrix
            if row is None: row = 0
            if col is None: col = 0
            origcol = col
            valarray = np.atleast_2d(np.array(vals, dtype=object))
            for rowvals in valarray:
                row += 1
                col = origcol
                for val in rowvals:
                    col += 1
                    try:
                        key = 'row:%s col:%s' % (row,col)
                        ws.cell(row=row, column=col, value=val)
                        if verbose: print('  Cell %s = %s' % (key, val))
                    except Exception as E:
                        errormsg = 'Could not write "%s" to %s: %s' % (val, key, repr(E))
                        raise Exception(errormsg)
        
        # Save
        wb.save(self.bytes)
        self.load()
        
        return None
        
    
    pass
    
    
def loadspreadsheet(filename=None, folder=None, fileobj=None, sheetname=None, sheetnum=None, asdataframe=None, header=True):
    '''
    Load a spreadsheet as a list of lists or as a dataframe. Read from either a filename or a file object.
    '''
    import xlrd # Optional import
    
    # Handle inputs
    if asdataframe is None: asdataframe = True
    if isinstance(filename, io.BytesIO): fileobj = filename # It's actually a fileobj
    if fileobj is None:
        fullpath = makefilepath(filename=filename, folder=folder)
        book = xlrd.open_workbook(fullpath)
    else:
        book = xlrd.open_workbook(file_contents=fileobj.read())
    if sheetname is not None: 
        sheet = book.sheet_by_name(sheetname)
    else:
        if sheetnum is None: sheetnum = 0
        sheet = book.sheet_by_index(sheetnum)
    
    # Load the raw data
    rawdata = []
    for rownum in range(sheet.nrows-header):
        rawdata.append(odict())
        for colnum in range(sheet.ncols):
            if header: attr = sheet.cell_value(0,colnum)
            else:      attr = 'Column %s' % colnum
            attr = ut.uniquename(attr, namelist=rawdata[rownum].keys(), style='(%d)')
            val = sheet.cell_value(rownum+header,colnum)
            try:
                val = float(val) # Convert it to a number if possible
            except: 
                try:    val = str(val)  # But give up easily and convert to a string (not Unicode)
                except: pass # Still no dice? Fine, we tried
            rawdata[rownum][str(attr)] = val
    
    # Convert to dataframe
    if asdataframe:
        cols = rawdata[0].keys()
        reformatted = []
        for oldrow in rawdata:
            newrow = list(oldrow[:])
            reformatted.append(newrow)
        dfdata = dataframe(cols=cols, data=reformatted)
        return dfdata
    
    # Or leave in the original format
    else:
        return rawdata



def savespreadsheet(filename=None, data=None, folder=None, sheetnames=None, close=True, formats=None, formatdata=None, verbose=False):
    '''
    Little function to format an output results nicely for Excel. Examples:
    
    import sciris as sc
    import pylab as pl
    
    # Simple example
    testdata1 = pl.rand(8,3)
    sc.savespreadsheet(filename='test1.xlsx', data=testdata1)
    
    # Include column headers
    test2headers = [['A','B','C']] # Need double to get right shape
    test2values = pl.rand(8,3).tolist()
    testdata2 = test2headers + test2values
    sc.savespreadsheet(filename='test2.xlsx', data=testdata2)
    
    # Multiple sheets
    testdata3 = [pl.rand(10,10), pl.rand(20,5)]
    sheetnames = ['Ten by ten', 'Twenty by five']
    sc.savespreadsheet(filename='test3.xlsx', data=testdata3, sheetnames=sheetnames)
    
    # Supply data as an odict
    testdata4 = sc.odict([('First sheet', pl.rand(6,2)), ('Second sheet', pl.rand(3,3))])
    sc.savespreadsheet(filename='test4.xlsx', data=testdata4, sheetnames=sheetnames)
    
    # Include formatting
    nrows = 15
    ncols = 3
    formats = {
        'header':{'bold':True, 'bg_color':'#3c7d3e', 'color':'#ffffff'},
        'plain': {},
        'big':   {'bg_color':'#ffcccc'}}
    testdata5  = pl.zeros((nrows+1, ncols), dtype=object) # Includes header row
    formatdata = pl.zeros((nrows+1, ncols), dtype=object) # Format data needs to be the same size
    testdata5[0,:] = ['A', 'B', 'C'] # Create header
    testdata5[1:,:] = pl.rand(nrows,ncols) # Create data
    formatdata[1:,:] = 'plain' # Format data
    formatdata[testdata5>0.7] = 'big' # Find "big" numbers and format them differently
    formatdata[0,:] = 'header' # Format header
    sc.savespreadsheet(filename='test5.xlsx', data=testdata5, formats=formats, formatdata=formatdata)
    '''
    import xlsxwriter # Optional import
    fullpath = makefilepath(filename=filename, folder=folder, default='default.xlsx')
    datadict   = odict()
    formatdict = odict()
    hasformats = (formats is not None) and (formatdata is not None)
    
    # Handle input arguments
    if isinstance(data, dict) and sheetnames is None:
        if verbose: print('Data is a dict, taking sheetnames from keys')
        sheetnames = data.keys()
        datadict   = odict(data) # Use directly, converting to odict first
        if hasformats: formatdict = odict(formatdata) #  NB, might be None but should be ok
    elif isinstance(data, dict) and sheetnames is not None:
        if verbose: print('Data is a dict, taking sheetnames from input')
        if len(sheetnames) != len(data):
            errormsg = 'If supplying data as a dict as well as sheetnames, length must matc (%s vs %s)' % (len(data), len(sheetnames))
            raise Exception(errormsg)
        datadict   = odict(data) # Use directly, but keep original sheet names
        if hasformats: formatdict = odict(formatdata)
    elif not isinstance(data, dict):
        if sheetnames is None:
            if verbose: print('Data is a simple array')
            sheetnames = ['Sheet1']
            datadict[sheetnames[0]]   = data # Set it explicitly
            formatdict[sheetnames[0]] = formatdata # Set it explicitly -- NB, might be None but should be ok
        else:
            if verbose: print('Data is a list, taking matching sheetnames from inputs')
            if len(sheetnames) == len(data):
                for s,sheetname in enumerate(sheetnames):
                    datadict[sheetname] = data[s] # Assume there's a 1-to-1 mapping
                    if hasformats: formatdict[sheetname] = formatdata[s]
            else:
                errormsg = 'Unable to figure out how to match %s sheet names with data of length %s' % (len(sheetnames), len(data))
                raise Exception(errormsg)
    else:
        raise Exception('Cannot figure out the format of the data') # This shouldn't happen!
        
    # Create workbook
    if verbose: print('Creating file %s' % fullpath)
    workbook = xlsxwriter.Workbook(fullpath)
    
    # Optionally add formats
    if formats is not None:
        if verbose: print('  Adding %s formats' % len(formats))
        workbook_formats = dict()
        for formatkey,formatval in formats.items():
            workbook_formats[formatkey] = workbook.add_format(formatval)
       
    # Actually write the data
    for sheetname in datadict.keys():
        if verbose: print('  Creating sheet %s' % sheetname)
        sheetdata   = datadict[sheetname]
        if hasformats: sheetformat = formatdict[sheetname]
        worksheet = workbook.add_worksheet(sheetname)
        for r,row_data in enumerate(sheetdata):
            if verbose: print('    Writing row %s/%s' % (r, len(sheetdata)))
            for c,cell_data in enumerate(row_data):
                if verbose: print('      Writing cell %s/%s' % (c, len(row_data)))
                if hasformats:
                    thisformat = workbook_formats[sheetformat[r][c]] # Get the actual format
                    worksheet.write(r, c, cell_data, thisformat) # Write with formatting
                else:
                    worksheet.write(r, c, cell_data) # Write without formatting
    
    # Either close the workbook and write to file, or return it for further working
    if close:
        if verbose: print('Saving file %s and closing' % filename)
        workbook.close()
        return None
    else:
        if verbose: print('Returning workbook')
        return workbook



##############################################################################
### Pickling support methods
##############################################################################

class Failed(object):
    ''' An empty class to represent a failed object loading '''
    failure_info = odict()
    def __init__(self, *args, **kwargs):
        pass

def makefailed(module_name=None, name=None, error=None):
    ''' Create a class -- not an object! -- that contains the failure info '''
    key = 'Failure %s' % (len(Failed.failure_info)+1)
    Failed.failure_info[key] = odict()
    Failed.failure_info[key]['module'] = module_name
    Failed.failure_info[key]['class'] = name
    Failed.failure_info[key]['error'] = repr(error)
    return Failed

class RobustUnpickler(pickle.Unpickler):
    ''' Try to import an object, and if that fails, return a Failed object rather than crashing '''
    def find_class(self, module_name, name, verbose=False):
        try:
            module = __import__(module_name)
            obj = getattr(module, name)
        except:
            try:
                string = 'from %s import %s as obj' % (module_name, name)
                exec(string)
            except Exception as E:
                if verbose: print('Unpickling warning: could not import %s.%s: %s' % (module_name, name, repr(E)))
                obj = makefailed(module_name=module_name, name=name, error=E)
        return obj

def unpickler(string=None, die=None):
    if die is None: die = False
    try: # Try pickle first
        obj = pkl.loads(string) # Actually load it -- main usage case
    except Exception as E:
        if die: 
            raise E
        else:
            try:    obj = dill.loads(string) # If that fails, try dill
            except: obj = RobustUnpickler(io.BytesIO(string)).load() # And if that trails, throw everything at it
    if isinstance(obj, Failed):
        print('Warning, the following errors were encountered during unpickling:')
        print(obj.failure_info)
    return obj

def savepickle(fileobj=None, obj=None):
        ''' Use pickle to do the salty work '''
        fileobj.write(pkl.dumps(obj, protocol=-1))
        return None
      
def savedill(fileobj=None, obj=None):
    ''' Use dill to do the sour work '''
    fileobj.write(dill.dumps(obj, protocol=-1))
    return None


##############################################################################
### Twisted pickling methods
##############################################################################

# NOTE: The code below is part of the Twisted package, and is included
# here to allow functools.partial() objects (among other things) to be 
# pickled; they are not for public consumption. --CK

# From: twisted/persisted/styles.py
# -*- test-case-name: twisted.test.test_persisted -*-
# Copyright (c) Twisted Matrix Laboratories.
# See LICENSE for details.

# Permission is hereby granted, free of charge, to any person obtaining
# a copy of this software and associated documentation files (the
# "Software"), to deal in the Software without restriction, including
# without limitation the rights to use, copy, modify, merge, publish,
# distribute, sublicense, and/or sell copies of the Software, and to
# permit persons to whom the Software is furnished to do so, subject to
# the following conditions:
#
# The above copyright notice and this permission notice shall be
# included in all copies or substantial portions of the Software.

if six.PY2:
    import cPickle
    class _UniversalPicklingError(pickle.PicklingError, cPickle.PicklingError):
        pass
else:
    _UniversalPicklingError = pickle.PicklingError

def pickleMethod(method):
    if six.PY3: return (unpickleMethod, (method.__name__,         method.__self__, method.__self__.__class__))
    else:       return (unpickleMethod, (method.im_func.__name__, method.im_self,  method.im_class))

def _methodFunction(classObject, methodName):
    methodObject = getattr(classObject, methodName)
    if six.PY3: return methodObject
    else:       return methodObject.im_func

def unpickleMethod(im_name, im_self, im_class):
    if im_self is None:
        return getattr(im_class, im_name)
    try:
        methodFunction = _methodFunction(im_class, im_name)
    except AttributeError:
        assert im_self is not None, "No recourse: no instance to guess from."
        if im_self.__class__ is im_class:
            raise
        return unpickleMethod(im_name, im_self, im_self.__class__)
    else:
        if six.PY3: maybeClass = ()
        else:       maybeClass = tuple([im_class])
        bound = types.MethodType(methodFunction, im_self, *maybeClass)
        return bound

cpreg.pickle(types.MethodType, pickleMethod, unpickleMethod)